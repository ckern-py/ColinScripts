#! python3
# WalkStatsAIO.py - Used to walk the user stat files from AHK located in C:\\Users\\Me\\Documents\\IT Service Desk\\Teamwork\\ITSDUtil Stats
# Will open each file, read it, gather the information, and present the information via excel files
# Make it only open each file once and gather all the needed info
# Creates 7 different excel files that present the information in different ways

import os
import re
import configparser
import openpyxl
import datetime
import time
import sys
import shutil
import collections

#if an error occurs, it should be displayed using the below function
def show_exception_and_exit(exc_type, exc_value, tb):
    import traceback
    traceback.print_exception(exc_type, exc_value, tb)
    input("Press enter to exit.")
    sys.exit(-1)

sys.excepthook = show_exception_and_exit
os.chdir('C:\\Users\\Me\\Documents\\IT Service Desk\\Teamwork\\ITSDUtil Stats')
formatDate = (datetime.date.today()).strftime('%Y%m%d')
start = time.time()
userStatsLocation = 'C:\\Users\\Me\\Documents\\Python\\UserStats\\'
os.makedirs(userStatsLocation + formatDate, exist_ok=True)

############################################################
##All regex to be compiled, i guess only need one?
namedStatFiles = re.compile(r'(\w+\d+)(-{1})(\d{8})(\.ini)') #user name #hypen separating name and date #date of the week #file extension, all .ini files 
############################################################
##arrays/dictionaries to be used. Defaults all new entries to 0 unless otherwise noted
foundUser = collections.defaultdict(lambda : 0)
foundUsers = collections.defaultdict(lambda : 0)
usesByKey = collections.defaultdict(lambda : 0)
allUsers = collections.defaultdict(lambda : 0)
allWeeks = collections.defaultdict(lambda : 0)
allSelections = collections.defaultdict(lambda : 0)
allWeeklyUsers = collections.defaultdict(lambda : 0)
############################################################
##parser initialization
iniFileParse = configparser.ConfigParser(strict=False, allow_no_value=True, delimiters=('='))
############################################################
##defined functions
##1########################################################
def files_Per_User(currUser):
	foundUser[currUser] += 1
##2#########################################################
def use_By_User(currUser):
	for allFoundSections in iniFileParse.sections():
		for allFoundItems in iniFileParse.items(allFoundSections):
			if allFoundItems[1] != '':
				foundUsers[currUser] += int(allFoundItems[1])
##3#########################################################
def use_By_Type():
	for allFoundSections in iniFileParse.sections():
		for allFoundItems in iniFileParse.items(allFoundSections):
			if allFoundItems[1] != '':
				usesByKey[allFoundSections] += int(allFoundItems[1])
##4#########################################################
def user_Use_By_Type(currUser):
	if not currUser in allUsers:
		allUsers[currUser] = collections.defaultdict(lambda : 0)
	for allFoundSections in iniFileParse.sections():
		for allFoundItems in iniFileParse.items(allFoundSections):
			if allFoundItems[1] != '':
				allUsers[currUser][allFoundSections] += int(allFoundItems[1])
##5#########################################################
def use_Per_Week(currWeek):
	if not currWeek in allWeeks:
		allWeeks[currWeek] = collections.defaultdict(lambda : 0)
	for allFoundSections in iniFileParse.sections():
		for allFoundItems in iniFileParse.items(allFoundSections):
			if allFoundItems[1] != '':
				allWeeks[currWeek][allFoundSections] += int(allFoundItems[1])	
##6#########################################################
def use_Per_Selection():
	for allFoundSections in iniFileParse.sections():
		if not allFoundSections in allSelections:
			allSelections[allFoundSections] = collections.defaultdict(lambda : 0)
		for allFoundItems in iniFileParse.items(allFoundSections):
			if allFoundItems[1] != '':
				allSelections[allFoundSections][allFoundItems[0]] += int(allFoundItems[1])
##7#########################################################
def user_Use_Per_Week(currWeek, currUser):
	if not currWeek in allWeeklyUsers:
		allWeeklyUsers[currWeek] = collections.defaultdict(lambda : 0)
	for allFoundSections in iniFileParse.sections():
		for allFoundItems in iniFileParse.items(allFoundSections):
			if allFoundItems[1] != '':
				allWeeklyUsers[currWeek][currUser] += int(allFoundItems[1])			
############################################################
def workbook_Create_New(sheetName):
	newestWorkBook = openpyxl.Workbook(write_only=True)
	newestWorkSheet = newestWorkBook.create_sheet()
	newestWorkSheet.title = sheetName
	return newestWorkBook, newestWorkSheet
############################################################
##actual code part? yes
print('Here we go...')
for statFiles in os.listdir(os.getcwd()):
	userFileSearch = namedStatFiles.search(statFiles)
	if userFileSearch is None:
		print('Skipping ' + statFiles)
		continue
	currUser = userFileSearch.group(1)
	currWeek = userFileSearch.group(3)
	files_Per_User(currUser)
	iniFileParse.read_file(open(statFiles, encoding='utf-16'))
	use_By_User(currUser)
	use_By_Type()
	user_Use_By_Type(currUser)
	use_Per_Week(currWeek)
	use_Per_Selection()
	user_Use_Per_Week(currWeek, currUser)
	iniFileParse.clear()
############################################################


##1##########################################################
#Below section gets all the users and finds the total number of weekly files for each
#print('Starting FilesPerUser') #def count_User_Files(statFiles):
newWorkBook, newWorkSheet = workbook_Create_New('FilesPerUser')
for a, b in foundUser.items():
	newWorkSheet.append([a.lower(), b])
newWorkBook.save(userStatsLocation  + formatDate + '\\' + formatDate + 'FilesPerUser.xlsx')
print('Saved ' + formatDate + 'FilesPerUser.xlsx')
###########################################################


##2##########################################################
##The below code gets the users and the total number of ITSDUtil uses to date
#print('Starting UseByUser')
newWorkBook, newWorkSheet = workbook_Create_New('UseByUser')
for c, d in foundUsers.items():
	newWorkSheet.append([c.lower(), d])
newWorkBook.save(userStatsLocation  + formatDate + '\\' + formatDate + 'UseByUser.xlsx')
print('Saved ' + formatDate + 'UseByUser.xlsx')
############################################################


##3#########################################################
##The below code will get all uses across the board and separate them by key([bracketwords])
#print('Starting UseByType')
newWorkBook, newWorkSheet = workbook_Create_New('UseByType')
for e, f in sorted(usesByKey.items()):
	newWorkSheet.append([e.lower(), f])
newWorkBook.save(userStatsLocation  + formatDate + '\\' + formatDate + 'UseByType.xlsx')
print('Saved ' + formatDate + 'UseByType.xlsx')
############################################################


##4##########################################################
#The below code will get uses by type for each user
#print('Starting UserUseByType')
maxKeys4 = 0
newWorkBook, newWorkSheet = workbook_Create_New('UserUseByType')
for key4 in sorted(allUsers.keys()):
	if len(allUsers[key4].keys()) > maxKeys4:
			maxKeys4 = len(allUsers[key4].keys())
			topKey4 = key4
			arrayOfKeys4 = []
			for aKeys4 in allUsers[key4].keys():
				arrayOfKeys4.append(aKeys4)
			arrayOfKeys4.sort(reverse=True)

arrayOfKeys4.insert(0, 'User')
newWorkSheet.append(arrayOfKeys4)
arrayOfKeys4.pop(0)

for currUser in sorted(allUsers.keys()):
	theCurrentUser = []
	theCurrentUser.insert(0, currUser)
	for u in range(0, maxKeys4):
		theCurrentUser.append((allUsers[currUser][arrayOfKeys4[u]]))
	newWorkSheet.append(theCurrentUser)

newWorkBook.save(userStatsLocation  + formatDate + '\\' + formatDate + 'UserUseByType.xlsx')
print('Saved ' + formatDate + 'UserUseByType.xlsx')
############################################################


#5##########################################################
#The below code will get uses per week across all users
#print('Starting UsePerWeek')	
#The following is for making a "Table" with Sources on top and Weeks down the side
maxKeys = 0
newWorkBook, newWorkSheet = workbook_Create_New('WeeklyUsage')
for key5 in sorted(allWeeks.keys()):
	if len(allWeeks[key5].keys()) > maxKeys:
			maxKeys = len(allWeeks[key5].keys())
			topKey = key5
			arrayOfKeys = []
			for aKeys in allWeeks[key5].keys():
				arrayOfKeys.append(aKeys)
			arrayOfKeys.sort(reverse=True)

arrayOfKeys.insert(0, 'Week')
newWorkSheet.append(arrayOfKeys)
arrayOfKeys.pop(0)

for monthDate in sorted(allWeeks.keys()):
	theCurrentWeek = []
	theCurrentWeek.insert(0, monthDate)
	for i in range(0, maxKeys):
		theCurrentWeek.append((allWeeks[monthDate][arrayOfKeys[i]]))
	newWorkSheet.append(theCurrentWeek)
		
newWorkBook.save(userStatsLocation  + formatDate + '\\' + formatDate + 'UsePerWeek.xlsx')
print('Saved ' + formatDate + 'UsePerWeek.xlsx')
###########################################################


#6##########################################################
#The below code will get uses per selection for each AHK item (The things in brackets[])
#print('Starting UsePerSelection')
workBook6 = openpyxl.Workbook(write_only=True)
for key6 in sorted(allSelections.keys()):
	workSheet6 = workBook6.create_sheet()
	workSheet6.title = key6
	for k, l in sorted(allSelections[key6].items()):
		workSheet6.append([k.lower(), l])
workBook6.save(userStatsLocation + formatDate + '\\' + formatDate + 'UsePerSelection.xlsx')
print('Saved ' + formatDate + 'UsePerSelection.xlsx')		
############################################################


##7#########################################################
#The below code will get use for each user, on a per week basis
#print('Starting UserUsePerWeek')	
newWorkBook, newWorkSheet = workbook_Create_New('FilesPerUser')
arrayOfWeeks = []
arrayOfUsers = []
for key7 in sorted(allWeeklyUsers.keys()):
	arrayOfWeeks.append(key7)
	for users in allWeeklyUsers[key7].keys():
		if not users in arrayOfUsers:
			arrayOfUsers.append(users.lower())

arrayOfUsers.sort()
arrayOfUsers.insert(0, 'Week')
newWorkSheet.append(arrayOfUsers)
arrayOfUsers.pop(0)

for currWeek in sorted(allWeeklyUsers.keys()):
	theCurrentWeekIs = []
	theCurrentWeekIs.insert(0, currWeek)
	for w in range(0, len(arrayOfUsers)):
		theCurrentWeekIs.append((allWeeklyUsers[currWeek][arrayOfUsers[w]]))
	newWorkSheet.append(theCurrentWeekIs)
	
totalArray = ['Total']
for t in range(0, len(arrayOfUsers)):
	finalFinding = 0
	for totals in allWeeklyUsers.keys():
		finalFinding += int(allWeeklyUsers[totals][arrayOfUsers[t]])
	totalArray.append(finalFinding)
newWorkSheet.append(totalArray)

newWorkBook.save(userStatsLocation + formatDate + '\\' + formatDate + 'UserUsePerWeek.xlsx')
print('Saved ' + formatDate + 'UserUsePerWeek.xlsx')
############################################################

print('Finished running all statistics in {0:0.2f} seconds'.format(time.time() - start))
print('Moving created files to LAN')
shutil.copytree(userStatsLocation + formatDate, '\\\\Removed\\name\\of\\server\\at\\work\\myfolder\\' + formatDate)
shutil.copytree(userStatsLocation + formatDate, '\\\\Removed\\name\\of\\server\\at\\work\\UserStats\\' + formatDate)
shutil.copytree(userStatsLocation + formatDate, '\\\\Removed\\name\\of\\server\\at\\work\\UserStats\\BackUpOfStats\\' + formatDate)

input("Press Enter to Close")
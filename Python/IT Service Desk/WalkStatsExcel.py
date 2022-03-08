#! python3
# WalkStatsExcel.py - Used to walk the user stat files from AHK located in C:\\Users\\Me\\Documents\\IT Service Desk\\Teamwork\\ITSDUtil Stats
# Will open each file, read it, gather the information, and present the information via excel files
# Creates 7 different excel files that present the information in different ways

def show_exception_and_exit(exc_type, exc_value, tb):
    import traceback
    traceback.print_exception(exc_type, exc_value, tb)
    input("Press enter to exit.")
    sys.exit(-1)


import os
import re
import configparser
import openpyxl
import datetime
import time
import sys
import shutil
sys.excepthook = show_exception_and_exit

os.chdir('C:\\Users\\Me\\Documents\\IT Service Desk\\Teamwork\\ITSDUtil Stats')

formatDate = (datetime.date.today()).strftime('%Y%m%d')
start = time.time()
userStatsLoc = 'C:\\Users\\Me\\Documents\\Python\\UserStats\\'
os.makedirs(userStatsLoc+ formatDate, exist_ok=True)

##1##########################################################
#Below section gets all the users and finds the total number of weekly files for each
print('Starting FilesPerUser')
userFiles = re.compile(r'''
	(\w+\d+)					#user name with hyphen(-)
	(-{1}\d{8})				    #date for week stat
	(\.ini)						#file extension, all .ini files
	''', re.VERBOSE)
foundUser = {}
for statFiles in os.listdir(os.getcwd()):
	mo = userFiles.search(statFiles)
	if mo == None:
		continue
	userName =mo.group(1)
	if userName in foundUser:
		countFiles = foundUser[userName]
		countFiles += 1
		foundUser[userName] = countFiles
	else:
		foundUser[userName] = 1
		
wb1 = openpyxl.Workbook(write_only=True)
ws1 = wb1.create_sheet()
ws1.title = 'FilesPerUser'
for a, b in sorted(foundUser.items()):
	ws1.append([a.lower(), b])
wb1.save(userStatsLoc  + formatDate + '\\' + formatDate + 'FilesPerUser.xlsx')
print('Saved ' + formatDate + 'FilesPerUser.xlsx')

print('Finished FilesPerUser\n')
###########################################################


##2##########################################################
##The below code gets the users and the total number of ITSDUtil uses to date
print('Starting UseByUser')
iniFileParse = configparser.ConfigParser(strict=False, allow_no_value=True, delimiters=('='))
userStatFiles = re.compile(r'''
	(\w+\d+)					#user name 
	(-{1}\d{8})				    #date for week stat with hyphen(-)
	(\.ini)						#file extension, all .ini files
	''', re.VERBOSE)
foundUsers = {}
for statFiles1 in os.listdir(os.getcwd()):
	mo1 = userStatFiles.search(statFiles1)
	if mo1 == None:
		continue
	userID =mo1.group(1)
	try: 
		iniFileParse.read_file(open(statFiles1, encoding='utf-16'))
	except PermissionError:
		print('PermissionsError: ' + statFiles1)
		continue
	for allFoundSections in iniFileParse.sections():
			for allFoundItems in iniFileParse.items(allFoundSections):
				if allFoundItems[1] == '':
					pass
				else:
					if userID in foundUsers:
						foundUsers[userID] += int(allFoundItems[1])
					else:
						foundUsers[userID] = int(allFoundItems[1])
	iniFileParse.clear()

wb2 = openpyxl.Workbook(write_only=True)
ws2 = wb2.create_sheet()
ws2.title = 'UseByUser'
for c, d in sorted(foundUsers.items()):
	ws2.append([c.lower(), d])
wb2.save(userStatsLoc  + formatDate + '\\' + formatDate + 'UseByUser.xlsx')
print('Saved ' + formatDate + 'UseByUser.xlsx')

print('Finished UseByUser\n')
############################################################


##3#########################################################
##The below code will get all uses across the board and separate them by key([bracketwords])
print('Starting UseByType')
config = configparser.ConfigParser(strict=False, allow_no_value=True, delimiters=('='))
usesByKey = {}
for statFiles2 in os.listdir(os.getcwd()):
	try:
		config.read_file(open(statFiles2, encoding='utf-16'))
		for foundSections in config.sections():
			for foundItems in config.items(foundSections):
				if foundItems[1] == '':
					pass
				else:
					if foundSections in usesByKey:
						countUse = usesByKey[foundSections]
						countUse += int(foundItems[1])
						usesByKey[foundSections] = countUse
					else:
						usesByKey[foundSections] = int(foundItems[1])
		config.clear()
	except PermissionError:
		print('Skipping ' + statFiles2)
		
wb3 = openpyxl.Workbook(write_only=True)
ws3 = wb3.create_sheet()
ws3.title = 'UseByType'
for e, f in sorted(usesByKey.items()):
	ws3.append([e.lower(), f])
wb3.save(userStatsLoc  + formatDate + '\\' + formatDate + 'UseByType.xlsx')
print('Saved ' + formatDate + 'UseByType.xlsx')

print('Finished UseByType\n')
############################################################


##4##########################################################
#The below code will get uses by type for each user
print('Starting UserUseByType')
config2 = configparser.ConfigParser(strict=False, allow_no_value=True, delimiters=('='))
allUsers = {}
userStatFiles = re.compile(r'''
	(\w+\d+)					#user name 
	(-{1}\d{8})				    #date for week stat with hyphen(-)
	(\.ini)						#file extension, all .ini files
	''', re.VERBOSE)
	
for statFiles3 in os.listdir(os.getcwd()):
	try:
		mo2 = userStatFiles.search(statFiles3)
		if mo2 == None:
			continue
		username2 = mo2.group(1)
		if username2 in allUsers:
			pass
		else:
			allUsers[username2]={}
		config2.read_file(open(statFiles3, encoding='utf-16'))
		for foundSections2 in config2.sections():
			for foundItems2 in config2.items(foundSections2):
				if foundItems2[1] == '':
					pass
				else:
					if foundSections2 in allUsers[username2]:
						countingUsage = allUsers[username2][foundSections2]
						countingUsage += int(foundItems2[1])
						allUsers[username2][foundSections2] = countingUsage
					else:
						allUsers[username2][foundSections2] = int(foundItems2[1])
		config2.clear()		
	except PermissionError:
		print('Skipping ' + statFiles3)
			

maxKeys4 = 0
wb4 = openpyxl.Workbook(write_only=True)
ws4 = wb4.create_sheet()
ws4.title = 'UserUseByType'

for key4 in sorted(allUsers.keys()):
	if len(allUsers[key4].keys()) > maxKeys4:
			maxKeys4 = len(allUsers[key4].keys())
			topKey4 = key4
			arrayOfKeys4 = []
			for aKeys4 in allUsers[key4].keys():
				arrayOfKeys4.append(aKeys4)
			arrayOfKeys4.sort(reverse=True)

arrayOfKeys4.insert(0, 'User')
ws4.append(arrayOfKeys4)
arrayOfKeys4.pop(0)

for currUser in sorted(allUsers.keys()):
	theCurrentUser = []
	theCurrentUser.insert(0, currUser)
	for u in range(0, maxKeys4):
		try:
			theCurrentUser.append((allUsers[currUser][arrayOfKeys4[u]]))
		except KeyError:
			theCurrentUser.append('')
	ws4.append(theCurrentUser)

wb4.save(userStatsLoc  + formatDate + '\\' + formatDate + 'UserUseByType.xlsx')
print('Saved ' + formatDate + 'UserUseByType.xlsx')
		
print('Finished UserUseByType\n')
############################################################


#5##########################################################
#The below code will get uses per week across all users
print('Starting UsePerWeek')
config3 = configparser.ConfigParser(strict=False, allow_no_value=True, delimiters=('='))
allWeeks = {}
userFilesStats = re.compile(r'''
	(\w+\d+-{1})			    #user name 
	(\d{8})						#date for week stat with hyphen(-)
	(\.ini)						#file extension, all .ini files
	''', re.VERBOSE)
	
for statFiles4 in os.listdir(os.getcwd()):
	try:
		mo3 = userFilesStats.search(statFiles4)
		if mo3 == None:
			continue
		weekDate = mo3.group(2)
		if weekDate in allWeeks:
			pass
		else:
			allWeeks[weekDate] = {}
		config3.read_file(open(statFiles4, encoding='utf-16'))
		for foundSections3 in config3.sections():
			for foundItems3 in config3.items(foundSections3):
				if foundItems3[1] == '':
					pass
				else:
					if foundSections3 in allWeeks[weekDate]:
						countingUse = allWeeks[weekDate][foundSections3]
						countingUse += int(foundItems3[1])
						allWeeks[weekDate][foundSections3] = countingUse
					else:
						allWeeks[weekDate][foundSections3] = int(foundItems3[1])
		config3.clear()		
	except PermissionError:
		print('Skipping ' + statFiles4)
			
#The following is for making a "Table" with Sources on top and Weeks down the side
maxKeys = 0
wb5 = openpyxl.Workbook(write_only=True)
ws5 = wb5.create_sheet()
ws5.title = 'WeeklyUsage'

for key5 in sorted(allWeeks.keys()):
	if len(allWeeks[key5].keys()) > maxKeys:
			maxKeys = len(allWeeks[key5].keys())
			topKey = key5
			arrayOfKeys = []
			for aKeys in allWeeks[key5].keys():
				arrayOfKeys.append(aKeys)
			#arrayOfKeys.sort()
			arrayOfKeys.sort(reverse=True)

arrayOfKeys.insert(0, 'Week')
ws5.append(arrayOfKeys)
arrayOfKeys.pop(0)

for monthDate in sorted(allWeeks.keys()):
	theCurrentWeek = []
	theCurrentWeek.insert(0, monthDate)
	for i in range(0, maxKeys):
		try:
			theCurrentWeek.append((allWeeks[monthDate][arrayOfKeys[i]]))
		except KeyError:
			theCurrentWeek.append('NA')
	ws5.append(theCurrentWeek)
		
wb5.save(userStatsLoc  + formatDate + '\\' + formatDate + 'UsePerWeek.xlsx')
print('Saved ' + formatDate + 'UsePerWeek.xlsx')
		
print('Finished UsePerWeek\n')
###########################################################


#6##########################################################
#The below code will get uses per selection for each AHK item (The things in brackets[])
print('Starting UsePerSelection')
config4 = configparser.ConfigParser(strict=False, allow_no_value=True, delimiters=('='))
allSelections = {}
for statFiles5 in os.listdir(os.getcwd()):
	try:
		config4.read_file(open(statFiles5, encoding='utf-16'))
		for foundSections4 in config4.sections():
			if foundSections4 in allSelections:
				pass
			else:
				allSelections[foundSections4] = {}
			for foundItems4 in config4.items(foundSections4):
				if foundItems4[1] == '':
					pass
				else:
					if foundItems4[0] in allSelections[foundSections4]:
						countingUse = allSelections[foundSections4][foundItems4[0]]
						countingUse += int(foundItems4[1])
						allSelections[foundSections4][foundItems4[0]] = countingUse
					else:
						allSelections[foundSections4][foundItems4[0]] = int(foundItems4[1])
		config4.clear()		
	except PermissionError:
		print('Skipping ' + statFiles5)
		

wb6 = openpyxl.Workbook(write_only=True)
for key6 in sorted(allSelections.keys()):
	ws6 = wb6.create_sheet()
	ws6.title = key6
	for k, l in sorted(allSelections[key6].items()):
		ws6.append([k.lower(), l])
wb6.save(userStatsLoc + formatDate + '\\' + formatDate + 'UsePerSelection.xlsx')
print('Saved ' + formatDate + 'UsePerSelection.xlsx')		
		
print('Finished UsePerSelection\n')
############################################################


##7#########################################################
#The below code will get use for each user, on a per week basis
print('Starting UserUsePerWeek')
config7 = configparser.ConfigParser(strict=False, allow_no_value=True, delimiters=('='))
allWeeklyUsers = {}
userWeeks = re.compile(r'''
	(\w+\d+)					#user name 
	(-{1})						#hypen separating name and date
	(\d{8})						#date of the week
	(\.ini)						#file extension, all .ini files
	''', re.VERBOSE)
	
for statFiles7 in os.listdir(os.getcwd()):
	try:
		mo7 = userWeeks.search(statFiles7)
		if mo7 == None:
			continue
		username7 = mo7.group(1)
		currWeek7 = mo7.group(3)
		if not currWeek7 in allWeeklyUsers:
			allWeeklyUsers[currWeek7]={}
		if not username7 in allWeeklyUsers[currWeek7]:
			allWeeklyUsers[currWeek7][username7] = 0
		config7.read_file(open(statFiles7, encoding='utf-16'))
		for foundSections7 in config7.sections():
			for foundItems7 in config7.items(foundSections7):
				if foundItems7[1] == '':
					pass
				else:
					countingUsage7 = allWeeklyUsers[currWeek7][username7]
					countingUsage7 += int(foundItems7[1])
					allWeeklyUsers[currWeek7][username7] = countingUsage7		
		config7.clear()		
	except PermissionError:
		print('Skipping ' + statFiles3)
			
wb7 = openpyxl.Workbook(write_only=True)
ws7 = wb7.create_sheet()
ws7.title = 'UserUsePerWeek'

arrayOfWeeks = []
arrayOfUsers = []
for key7 in sorted(allWeeklyUsers.keys()):
	arrayOfWeeks.append(key7)
	for users in allWeeklyUsers[key7].keys():
		if not users in arrayOfUsers:
			arrayOfUsers.append(users.lower())

arrayOfUsers.sort()
arrayOfUsers.insert(0, 'Week')
ws7.append(arrayOfUsers)
arrayOfUsers.pop(0)

for currWeek in sorted(allWeeklyUsers.keys()):
	theCurrentWeekIs = []
	theCurrentWeekIs.insert(0, currWeek)
	for w in range(0, len(arrayOfUsers)):
		try:
			theCurrentWeekIs.append((allWeeklyUsers[currWeek][arrayOfUsers[w]]))
		except KeyError:
			theCurrentWeekIs.append('')
	ws7.append(theCurrentWeekIs)
	
totalArray = ['Total']
for t in range(0, len(arrayOfUsers)):
	finalFinding = 0
	for totals in allWeeklyUsers.keys():
		try: 
			finalFinding += int(allWeeklyUsers[totals][arrayOfUsers[t]])
		except KeyError:
			pass
	totalArray.append(finalFinding)
ws7.append(totalArray)

wb7.save(userStatsLoc + formatDate + '\\' + formatDate + 'UserUsePerWeek.xlsx')
print('Saved ' + formatDate + 'UserUsePerWeek.xlsx')
		
print('Finished UserUsePerWeek\n')
############################################################

print('Finished running all statistics in {0:0.2f} seconds'.format(time.time() - start))

print('Moving created files to LAN')
shutil.copytree(userStatsLoc + formatDate, '\\\\Removed\\name\\of\\server\\at\\work\\myfolder\\' + formatDate)
shutil.copytree(userStatsLoc + formatDate, '\\\\Removed\\name\\of\\server\\at\\work\\UserStats\\' + formatDate)
shutil.copytree(userStatsLoc + formatDate, '\\\\Removed\\name\\of\\server\\at\\work\\UserStats\\BackUpOfStats\\' + formatDate)

input("Press Enter to Close")